
from builtins import isinstance
import csv
from operator import attrgetter
from os.path import splitext
from collections import OrderedDict
import time

from django.utils.html import strip_tags
from openpyxl.reader.excel import load_workbook
try:
    from xlsxwriter import Workbook
except ImportError:
    pass

from vavilov.models import Trait, Plant

PLANT_ID = 'plant_name'
TRAIT_HEADER_NAME = 'Caracteristica'
TRAIT_TYPE_NAME = 'Tipo'
VALUE_HEARDER_NAME = 'Valor'
DATE_HEADER_NAME = 'Fecha'
USER_HEDER_NAME = 'Autor'

MANDATORY_COLS = [PLANT_ID, TRAIT_HEADER_NAME, TRAIT_TYPE_NAME,
                  VALUE_HEARDER_NAME, DATE_HEADER_NAME, USER_HEDER_NAME]


class trt_dialect(csv.excel):
    delimiter = ','
    skipinitialspace = True
    doublequote = False
    lineterminator = '\n'


def _parse_traits(fpath):
    ext = splitext(fpath)[1]
    if ext in ('.trt', 'csv'):
        trait_names = _parse_csv_traits(fpath)
    elif ext in ('.xlsx', '.xlsm'):
        trait_names = _parse_excel_traits(fpath)
    else:
        raise RuntimeError('Not known file format')

    traits = []
    for trait_name in trait_names:
        traits.append(Trait.objects.get(name=trait_name))

    return sorted(traits, key=attrgetter('name'))


def _parse_csv_traits(fpath):
    for trait_line in csv.DictReader(open(fpath, "r"), dialect=trt_dialect):
        yield trait_line['trait']


def _parse_excel_traits(fpath):
    for row in excel_dict_reader(fpath):
        yield row['name']


def write_row_in_sheet(sheet, row_content, row_index):
    for col_idx, column in enumerate(row_content):
        value = column["value"]
        format_ = column["format"]
        sheet.write(row_index, col_idx, value, format_)


def write_excel_observations_skeleton(entries_fpath, traits_fpath,
                                      out_fhand, accession_header=None,
                                      synonym_header=None, row_header=None,
                                      column_header=None,
                                      pot_number_header=None,
                                      vba_macro=None,
                                      rows_per_plant=1):

    utc_offset = int(-(time.timezone / 3600))
    sign = '+' if utc_offset >= 0 else '-'

    relative_hour = '{0}{1:02d}00'.format(sign, utc_offset)
    columns = []
    if accession_header is not None:
        columns.append(accession_header)
    if synonym_header is not None:
        columns.append(synonym_header)
    if row_header is not None:
        columns.append(row_header)
    if column_header is not None:
        columns.append(column_header)
    if pot_number_header is not None:
        columns.append(pot_number_header)

    columns += MANDATORY_COLS
    traits = _parse_traits(traits_fpath)

    workbook = Workbook(out_fhand.name)

    workbook.set_vba_name('ThisWorkbook')

    sheet = workbook.add_worksheet()
    sheet.set_vba_name('Hoja1')

    if vba_macro:
        workbook.add_vba_project(vba_macro)

    # header
    locked = workbook.add_format()
    locked.set_locked(True)

    unlocked = workbook.add_format()
    unlocked.set_locked(False)

    date_fmt = "yyyy-mm-dd hh:mm:ss\"{}\"".format(relative_hour)
    date_format = workbook.add_format({'num_format': date_fmt,
                                       'locked': 1})

    # sheet.protect()
    max_lengths = {}
    for idx, col in enumerate(columns):
        max_lengths[idx] = [len(col)]
        sheet.write(0, idx, col, locked)

    # Line per Maceta
    row_index = 1
    for plant_def in csv.DictReader(open(entries_fpath, "r")):
        plant = Plant.objects.get(plant_name=plant_def['unique_id'])
        for trait in traits:
            row_content = []
            for column_index, column in enumerate(columns):
                if column == PLANT_ID:
                    value = plant.plant_name
                    format_ = locked
                elif accession_header and column == accession_header:
                    value = plant.accession.accession_number
                    format_ = locked
                elif synonym_header and column == synonym_header:
                    try:
                        value = plant.accession.collecting_accession[-1]
                    except TypeError:
                        value = None
                    format_ = locked
                elif row_header and column == row_header:
                    value = plant.row
                    format_ = locked
                elif column_header and column == column_header:
                    value = plant.column
                    format_ = locked
                elif pot_number_header and column == pot_number_header:
                    value = plant.pot_number
                    format_ = locked
                elif column == TRAIT_HEADER_NAME:
                    value = trait.name
                    format_ = locked
                elif column == TRAIT_TYPE_NAME:
                    value = trait.type.name
                    format_ = locked
                elif column == DATE_HEADER_NAME:
                    value = None
                    format_ = date_format
                else:
                    value = None
                    format_ = unlocked
                row_content.append({'value': value, 'format': format_})
            for _ in range(rows_per_plant):
                write_row_in_sheet(sheet, row_content, row_index)
                row_index += 1
            len_value = 0 if value is None else len(value)
            max_lengths[column_index].append(len_value)

    fecha_col_index = columns.index(DATE_HEADER_NAME)
    valor_col_index = columns.index(VALUE_HEARDER_NAME)
    autor_col_index = columns.index(USER_HEDER_NAME)
    for col_index, lengths in max_lengths.items():
        if col_index in (fecha_col_index, valor_col_index, autor_col_index):
            length = 25
        else:
            length = max(lengths)
        sheet.set_column(col_index, col_index, length + 2)

    workbook.close()


def queryset_to_row(queryset, table_class):
    entries = table_class(queryset)

    header = [field.verbose_name for field in entries.base_columns.values()]
    yield(header)
    for row in entries.rows:
        row_cells = []
        for _, cell in row.items():
            if isinstance(cell, str):
                cell = strip_tags(cell)
            else:
                cell = str(cell)
            row_cells.append(cell)
        yield row_cells


def create_excel_from_queryset(out_fhand, queryset, table, in_memory=False):
    workbook = Workbook(out_fhand, {'in_memory': in_memory})

    worksheet = workbook.add_worksheet()
    max_lengths = {}
    for idx, row in enumerate(queryset_to_row(queryset, table)):
        for col_idx, cell in enumerate(row):
            if col_idx not in max_lengths:
                max_lengths[col_idx] = []
            max_lengths[col_idx].append(len(cell))
        worksheet.write_row(idx, 0, row)
    for col_index, lengths in max_lengths.items():
        length = max(lengths)
        worksheet.set_column(col_index, col_index, length + 2)
    workbook.close()
    return out_fhand


def get_sheet_names(fpath):
    wb = load_workbook(fpath, read_only=True)
    return wb.get_sheet_names()


def excel_dict_reader(fpath, data_only=True, sheet_name=None):
    wb = load_workbook(fpath, read_only=True, data_only=data_only)
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
    header_pos = OrderedDict()
    first = True
    for row in sheet.rows:
        if first:
            for index, cell in enumerate(row):
                header = cell.value
                if not header:
                    continue
                try:
                    header = header.strip()
                except Exception:
                    pass
                header_pos[header] = index
            first = False
            continue
        row_dict = OrderedDict()
        for header, pos in header_pos.items():
            value = row[pos].value
            try:
                value = value.strip()
            except Exception:
                pass
            row_dict[header] = value
        if any(row_dict.values()):
            yield row_dict
        else:
            break
