import csv
import os
from random import Random

from django.contrib.auth.models import User, Group
from django.db import transaction
from guardian.shortcuts import assign_perm
from openpyxl.reader.excel import load_workbook

from vavilov.db_management.base import comma_dialect
from vavilov.models import (PlantPart, Observation, Trait, Assay, AssayPlant,
                            Cvterm, TraitProp, AssayTrait, AssayProp, Plant,
                            Accession)

TRAIT_PROPS_CV = 'trait_props'
TRAIT_TYPES_CV = 'trait_types'


def add_or_load_assays(fpath):
    fhand = open(fpath)
    with transaction.atomic():
        for entry in csv.DictReader(fhand, dialect=comma_dialect):
            assay_data = {}
            props = {}
            for key, value in entry.items():
                if not value:
                    continue
                if key in ('name', 'description', 'start_date', 'end_date',
                           'location'):
                    assay_data[key] = value
                elif key == 'owner':
                    owner = value
                else:
                    props[key] = value

            try:
                owner = User.objects.get(username=owner)
            except User.DoesNotExist:
                msg = 'user {} owner of assay {} does not exist'
                raise RuntimeError(msg.format(owner, entry['name']))

            assay, created = Assay.objects.get_or_create(**assay_data,
                                                         owner=owner)

            if created:
                for key, value in props.items():
                    type_ = Cvterm.objects.get(cv__name='assay_props',
                                               name=key)
                    AssayProp.objects.create(assay=assay, type=type_,
                                             value=value)
                group, created = Group.objects.get_or_create(name=assay.name)
                assign_perm('view_assay', group, assay)
                group.user_set.add(owner)


def add_or_load_observation(plant_part, trait, assay, value, creation_time,
                            observer=None):
    try:
        trait = Trait.objects.get(name=trait, assaytrait__assay=assay)
    except Trait.DoesNotExist:
        raise ValueError('Trait not loaded yet in db: {}'.format(trait))
    plant = plant_part.plant
    try:
        AssayPlant.objects.get(assay=assay, plant=plant)
    except AssayPlant.DoesNotExist:
        msg = 'This assay {} and this plant {} are not related'
        raise RuntimeError(msg.format(assay, plant))

    obs = Observation.objects.get_or_create(plant_part=plant_part,
                                            trait=trait, assay=assay,
                                            value=value,
                                            creation_time=creation_time,
                                            observer=observer)[0]
    return obs


def add_or_load_excel_traits(fpath, assays):
    wb = load_workbook(fpath)
    sheet = wb.active
    header_pos = {}
    first = True
    for row in sheet.iter_rows():
        if first:
            for index, cell in enumerate(row):
                header = cell.value
                if not header:
                    continue
                header_pos[header] = index
            first = False
            continue

        name = row[header_pos['name']].value
        type_ = row[header_pos['type']].value
        description = row[header_pos['description']].value
        try:
            type_ = Cvterm.objects.get(name=type_, cv__name=TRAIT_TYPES_CV)
        except Cvterm.DoesNotExist:
            msg = 'Trait type not loaded yet in db: {}'.format(type_)
            raise RuntimeError(msg)

        trait, trait_created = Trait.objects.get_or_create(name=name, type=type_)
        for assay in assays:
            try:
                assay = Assay.objects.get(name=assay)
            except Assay.DoesNotExist:
                raise RuntimeError('Assay not loaded yet in db: {}'.format(assay))

            created = AssayTrait.objects.get_or_create(assay=assay,
                                                       trait=trait)[1]
            if created:
                assign_perm('view_trait', assay.owner, trait)

        if trait_created and description:
            desc_type = Cvterm.objects.get(cv__name=TRAIT_PROPS_CV,
                                           name='description')
            TraitProp.objects.create(trait=trait, type=desc_type,
                                     value=description)


def add_or_load_plants(fpath, assay, experimental_field_header=None,
                       row_header=None, column_header=None,
                       pot_number_header=None, view_perm_group=None):
    fhand = open(fpath)
    assay = Assay.objects.get(name=assay)
    if view_perm_group is None:
        view_perm_group = assay.name
    group = Group.objects.get(name=view_perm_group)
    with transaction.atomic():
        for entry in csv.DictReader(fhand, dialect=comma_dialect):
            plant_name = entry['plant_id']
            accession_code = entry['accession']

            exp_field = entry.get(experimental_field_header, None)
            row = entry.get(row_header, None)
            column = entry.get(column_header, None)
            pot_number = entry.get(pot_number_header, None)

            new_plant = False
            try:
                plant = Plant.objects.get(plant_name=plant_name)
            except Plant.DoesNotExist:
                new_plant = True

            if new_plant:
                accession = Accession.objects.get(accession_number=accession_code)

                plant = Plant.objects.create(plant_name=plant_name,
                                             accession=accession,
                                             experimental_field=exp_field,
                                             row=row, column=column,
                                             pot_number=pot_number)
                assign_perm('view_plant', group, plant)
            AssayPlant.objects.get_or_create(assay=assay, plant=plant)


class RandomNameSequence:
    """An instance of _RandomNameSequence generates an endless
    sequence of unpredictable strings which can safely be incorporated
    into file names.  Each string is six characters long.  Multiple
    threads can safely use the same instance at the same time.

    _RandomNameSequence is an iterator."""

    characters = "abcdefghijklmnopqrstuvwxyz0123456789"

    @property
    def rng(self):
        cur_pid = os.getpid()
        if cur_pid != getattr(self, '_rng_pid', None):
            self._rng = Random()
            self._rng_pid = cur_pid
        return self._rng

    def __iter__(self):
        return self

    def __next__(self):
        c = self.characters
        choose = self.rng.choice
        letters = [choose(c) for dummy in range(8)]
        return ''.join(letters)

NAMER = RandomNameSequence()


def suggest_plant_part_uid(plant_name, plant_part):
    if plant_part == 'plant':
        return '{}_{}'.format(plant_name, plant_part)
    part_uid = '{}_{}_{}'.format(plant_name, plant_part, next(NAMER))
    try:
        PlantPart.objects.get(plant_part_uid=part_uid)
    except PlantPart.DoesNotExist:
        return part_uid
    else:
        return suggest_plant_part_uid(plant_name, plant_part)
