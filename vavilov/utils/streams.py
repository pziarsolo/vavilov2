import csv
import logging
from time import time
from django.http.response import StreamingHttpResponse, HttpResponse

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

from vavilov.conf.settings import MAX_OBS_TO_EXCEL, APP_LOGGER
from vavilov.utils.column_format import queryset_to_columns
logger = logging.getLogger(APP_LOGGER)


def calc_duration(action, prev_time):
    now = time()
    logger.debug('{}: Took {} secs'.format(action, round(now - prev_time, 2)))
    return now


class Echo(object):
    """An object that implements just the write method of the file-like
    interface.
    """

    def write(self, value):
        """Write the value by returning it, instead of storing in a buffer."""
        return value


MSG = """<script>
    alert('To much observations( more than {}). Please contact with the admin to obtain the excel file :)');
    window.history.back();
</script>""".format(MAX_OBS_TO_EXCEL)


def return_csv_response(queryset, table_class, column_format=False):
    prev_time = time()
    if queryset.count() > MAX_OBS_TO_EXCEL:
        return HttpResponse(MSG, content_type="text/html")
    prev_time = calc_duration('Csv Query Count', prev_time)
    rows = table_class(queryset).as_values()
    if column_format:
        rows = queryset_to_columns(rows)
    prev_time = calc_duration('table to rows', prev_time)
    pseudo_buffer = Echo()
    writer = csv.writer(pseudo_buffer, delimiter=',', quoting=csv.QUOTE_MINIMAL)
    response = StreamingHttpResponse((writer.writerow(row) for row in rows),
                                     content_type="text/csv")
    response['Content-Disposition'] = 'attachment; filename="search_result.csv"'
    return response


def return_excel_response(queryset, table_class, column_length=None):
    if queryset.count() > MAX_OBS_TO_EXCEL:
        return HttpResponse(MSG, content_type="text/html")

    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=content_type)
    rows = table_class(queryset).as_values()
    wb = create_workbook_from_queryset(rows, column_length)
    response['Content-Disposition'] = 'attachment; filename="somefilename.xlsx"'
    wb.save(response)
    return response


def create_excel_from_queryset(out_fpath, queryset, table, column_length=None):
    rows = table(queryset).as_values()
    wb = create_workbook_from_queryset(rows, column_length=column_length)
    wb.save(out_fpath)


def create_workbook_from_queryset(rows, column_length=None):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    if column_length is None:
        column_length = 13
    first = True
    for row in rows:
        if first:
            first = False
            for col_index in range(len(row)):
                column = get_column_letter(col_index + 1)
                ws.column_dimensions[column].width = column_length
        ws.append(row)
    return wb
