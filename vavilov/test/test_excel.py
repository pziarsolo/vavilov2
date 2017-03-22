from os.path import join
import tempfile

from django.test import TestCase
from openpyxl import load_workbook

from vavilov.db_management.base import (load_initial_data, INITIAL_DATA_DIR,
                                        add_or_load_persons, add_accessions)
from vavilov.db_management.excel import (write_excel_observations_skeleton,)
from vavilov.db_management.fieldbook import (OUR_TIMEZONE)
from vavilov.db_management.phenotype import (add_or_load_excel_observations,
                                             add_or_load_excel_traits,
                                             add_or_load_assays,
                                             add_or_load_plants)
from vavilov.db_management.tests import (TEST_DATA_DIR, create_test_users,
                                         load_test_data)
from vavilov.models import Trait, Observation

TRAITS_FIELDBOOK = join(TEST_DATA_DIR, 'traits.fieldbook.trt')
PLANT_FIELDBOOK = join(TEST_DATA_DIR, 'plant_fields.csv')


class ExcelLoadTests(TestCase):

    def setUp(self):
        load_initial_data()
        create_test_users()
        add_or_load_persons(open(join(INITIAL_DATA_DIR, 'vavilov_person.csv')))
        add_accessions(open(join(TEST_DATA_DIR, 'accessions.csv')))
        add_or_load_assays(join(TEST_DATA_DIR, 'assays.csv'))
        add_or_load_plants(join(TEST_DATA_DIR, 'plants.csv'), assay='NSF1')

    def test_add_excel_traits(self):
        add_or_load_excel_traits(join(TEST_DATA_DIR, 'traits.xlsx'),
                                 assays=['NSF1', 'NSF2'])

        assert Trait.objects.all().count() == 3

    def test_add_excel_observations(self):
        add_or_load_excel_traits(join(TEST_DATA_DIR, 'traits.xlsx'),
                                 assays=['NSF1', 'NSF2'])
        assert len(Observation.objects.all()) == 0
        add_or_load_excel_observations(join(TEST_DATA_DIR, 'observations1.xlsx'))

        obs = Observation.objects.all()
        assert obs.count() == 12
        creation_time = obs[0].creation_time
        assert creation_time.tzname() == 'UTC'
        assert str(creation_time).startswith("2016-06-07 11:34:")

        creation_time = OUR_TIMEZONE.normalize(creation_time.astimezone(OUR_TIMEZONE))
        assert creation_time.tzname() == 'CEST'
        assert str(creation_time).startswith("2016-06-07 13:34:")

        add_or_load_excel_observations(join(TEST_DATA_DIR, 'observations1.xlsx'))
        obs = Observation.objects.all()
        assert obs.count() == 12


class ExcelCreateTest(TestCase):

    def setUp(self):
        load_test_data()

    def test_create_excel(self):
        # creates excel
        out_fhand = tempfile.NamedTemporaryFile(suffix='.xlsm')
        write_excel_observations_skeleton(PLANT_FIELDBOOK,
                                          TRAITS_FIELDBOOK,
                                          out_fhand,
                                          accession_header='BGV',
                                          synonym_header='Accession',
                                          row_header='Fila_F',
                                          pot_number_header='Maceta_M',
                                          rows_per_plant=2)

        workbook = load_workbook(out_fhand.name, read_only=True)
        ws = workbook.active

        # Check header
        assert ws['A1'].value == "BGV", "Incorrect header value: %s" % (ws['C1'].value)
        assert ws['B1'].value == "Accession", "Incorrect header value: %s" % (ws['B1'].value)
        assert ws['C1'].value == "Fila_F", "Incorrect header value: %s" % (ws['C1'].value)
        assert ws['D1'].value == "Maceta_M", "Incorrect header value: %s" % (ws['D1'].value)
        assert ws['E1'].value == "plant_name", "Incorrect header value: %s" % (ws['E1'].value)
        assert ws['F1'].value == "Caracteristica", "Incorrect header value: %s" % (ws['F1'].value)
        assert ws['G1'].value == "Tipo", "Incorrect header value: %s" % (ws['G1'].value)
        assert ws['H1'].value == "Valor", "Incorrect header value: %s" % (ws['H1'].value)
        assert ws['I1'].value == "Fecha", "Incorrect header value: %s" % (ws['I1'].value)
        assert ws['J1'].value == "Autor", "Incorrect header value: %s" % (ws['J1'].value)
        assert ws['A2'].value == "BGV000917", "Incorrect plant accession: %s" % (ws['A2'].value)
        assert ws['A21'].value == "BGV000934", "Incorrect plant accession: %s" % (ws['A21'].value)
        assert ws['A22'].value == "BGV000935", "Incorrect plant accession: %s" % (ws['A22'].value)
        assert ws['A24'].value == "BGV000935", "Incorrect plant accession: %s" % (ws['A24'].value)

    def test_create_excel2(self):
        # creates excel
        out_fhand = tempfile.NamedTemporaryFile(suffix='.xlsm')
        write_excel_observations_skeleton(PLANT_FIELDBOOK,
                                          TRAITS_FIELDBOOK,
                                          out_fhand,
                                          accession_header='BGV',
                                          synonym_header='Accession',
                                          row_header='Fila_F',
                                          pot_number_header='Maceta_M',
                                          rows_per_plant=2)

        workbook = load_workbook(out_fhand.name, read_only=True)
        ws = workbook.active

        # Check header
        assert ws['A1'].value == "BGV", "Incorrect header value: %s" % (ws['C1'].value)
        assert ws['B1'].value == "Accession", "Incorrect header value: %s" % (ws['B1'].value)
        assert ws['C1'].value == "Fila_F", "Incorrect header value: %s" % (ws['C1'].value)
        assert ws['D1'].value == "Maceta_M", "Incorrect header value: %s" % (ws['D1'].value)
        assert ws['E1'].value == "plant_name", "Incorrect header value: %s" % (ws['E1'].value)
        assert ws['F1'].value == "Caracteristica", "Incorrect header value: %s" % (ws['F1'].value)
        assert ws['G1'].value == "Tipo", "Incorrect header value: %s" % (ws['G1'].value)
        assert ws['H1'].value == "Valor", "Incorrect header value: %s" % (ws['H1'].value)
        assert ws['I1'].value == "Fecha", "Incorrect header value: %s" % (ws['I1'].value)
        assert ws['J1'].value == "Autor", "Incorrect header value: %s" % (ws['J1'].value)
        assert ws['A21'].value == "BGV000934", "Incorrect plant accession: %s" % (ws['A21'].value)
        assert ws['A22'].value == "BGV000935", "Incorrect plant accession: %s" % (ws['A22'].value)
        assert ws['A24'].value == "BGV000935", "Incorrect plant accession: %s" % (ws['A24'].value)
